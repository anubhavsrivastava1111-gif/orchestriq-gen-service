"""
OrchestrIQ Document Intelligence Engine v4.1 — Blueprint Engine
Generic spec-driven Excel renderer. The AI (or a deterministic fallback)
produces a JSON blueprint describing sheets, columns, data-generation rules,
formula columns, group summaries, and dashboard KPIs. This module renders
ANY blueprint into a formatted workbook with native Excel formulas.

Blueprint shape:
{
 "title": "...",
 "sheets": [
  {"name":"Assumptions","type":"kv","rows":[["label", value, "note"], ...]},
  {"name":"Employee Data","type":"table","row_count":150,
   "columns":[
     {"h":"Employee ID","gen":{"kind":"id","prefix":"EMP-","start":1001}},
     {"h":"Employee Name","gen":{"kind":"name"}},
     {"h":"Department","gen":{"kind":"choice","values":["Ops","Finance"]}},
     {"h":"Scheduled Hours","gen":{"kind":"number","min":168,"max":184,"decimals":0}},
     {"h":"Utilization %","gen":{"kind":"number","min":0.7,"max":0.95,"decimals":3,"format":"percent"}},
     {"h":"Available Hours","formula":"{Scheduled Hours}-{Leave Hours}","format":"number"},
   ]},
  {"name":"Department Summary","type":"summary","source":"Employee Data","group_by":"Department",
   "aggregates":[{"h":"Headcount","kind":"count"},
                 {"h":"Total FTE","kind":"sum","col":"FTE"},
                 {"h":"Avg Utilization","kind":"avg","col":"Utilization %","format":"percent"}]},
  {"name":"Dashboard","type":"dashboard",
   "kpis":[{"label":"Total Employees","ref":{"sheet":"Employee Data","agg":"count","col":"Employee ID"}},
           {"label":"Avg Utilization","ref":{"sheet":"Employee Data","agg":"avg","col":"Utilization %"},"format":"percent"}],
   "charts":[{"title":"FTE by Department","type":"bar","source":"Department Summary","cat_col":"Department","val_col":"Total FTE"}]}
 ]
}
"""
import io
import random
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.table import Table, TableStyleInfo

NAVY = "1E3A5F"; TEAL = "14B8A6"; LIGHT = "F1F5F9"; WHITE = "FFFFFF"
GREY = "64748B"; GREEN = "16A34A"; RED = "DC2626"
_thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

FIRST = ["Aarav", "Vivaan", "Aditya", "Arjun", "Sai", "Reyansh", "Krishna", "Ishaan",
         "Rohan", "Kabir", "Ananya", "Diya", "Aadhya", "Saanvi", "Pari", "Anika",
         "Navya", "Riya", "Myra", "Ira", "Rahul", "Priya", "Amit", "Sneha", "Vikram",
         "Pooja", "Rajesh", "Kavya", "Suresh", "Meera", "James", "Maria", "David",
         "Sarah", "Michael", "Emma", "Daniel", "Olivia", "John", "Sophia"]
LAST = ["Sharma", "Verma", "Patel", "Gupta", "Singh", "Kumar", "Reddy", "Nair",
        "Iyer", "Das", "Khan", "Mehta", "Joshi", "Rao", "Chopra", "Mishra",
        "Agarwal", "Bose", "Pillai", "Kapoor", "Smith", "Johnson", "Garcia",
        "Brown", "Miller", "Davis", "Wilson", "Anderson", "Thomas", "Martin"]

FORMATS = {"currency": None, "percent": "0.0%", "number": "#,##0",
           "decimal": "#,##0.00", "text": "@", "hours": "#,##0.0"}


def _fmt_code(fmt, sym):
    if fmt == "currency":
        return f'"{sym}"#,##0'
    return FORMATS.get(fmt, "General")


def _hdr_cell(c):
    c.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER


def _title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=15, color=NAVY, name="Calibri")
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(size=9, color=GREY, italic=True)


class _Gen:
    """Deterministic-ish data synthesizer per gen rule."""
    def __init__(self, seed=42):
        self.rng = random.Random(seed)
        self._name_i = 0

    def value(self, rule, row_idx, ctx, salt=0):
        k = rule.get("kind", "number")
        if k == "id":
            return f"{rule.get('prefix','ID-')}{rule.get('start',1000)+row_idx}"
        if k == "name":
            f = FIRST[(row_idx * 7 + 3 + salt * 13) % len(FIRST)]
            l = LAST[(row_idx * 11 + 5 + salt * 17) % len(LAST)]
            return f"{f} {l}"
        if k == "choice":
            vals = rule.get("values") or ["A", "B"]
            if rule.get("sequential"):
                return vals[row_idx % len(vals)]
            return self.rng.choice(vals)
        if k == "choice_dependent":
            # value depends on another column's value: {"map":{"Ops":["Team A","Team B"]}}
            dep = ctx.get(rule.get("on", ""))
            opts = (rule.get("map") or {}).get(dep) or rule.get("default") or ["Team 1"]
            return self.rng.choice(opts)
        if k == "number":
            lo, hi = float(rule.get("min", 0)), float(rule.get("max", 100))
            v = self.rng.uniform(lo, hi)
            d = int(rule.get("decimals", 0))
            return round(v, d) if d > 0 else int(round(v))
        if k == "const":
            return rule.get("value", "")
        return ""


def _translate_formula(tmpl, colmap, row):
    """Replace {Column Name} tokens with cell refs for this row."""
    def rep(m):
        name = m.group(1)
        ci = colmap.get(name)
        if ci is None:
            return "0"
        return f"{get_column_letter(ci)}{row}"
    return "=" + re.sub(r"\{([^}]+)\}", rep, tmpl.lstrip("="))


def render_blueprint(bp: dict, currency_symbol: str = "\u20b9") -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    gen = _Gen()
    sheet_meta = {}  # name -> {"colmap":..., "n_rows":..., "start_row":...}

    sheets = bp.get("sheets") or []
    # Render non-dashboard sheets first so dashboards can reference them
    ordered = [s for s in sheets if s.get("type") != "dashboard"] + \
              [s for s in sheets if s.get("type") == "dashboard"]

    for spec in ordered:
        st = spec.get("type", "table")
        name = str(spec.get("name", "Sheet"))[:28]
        ws = wb.create_sheet(name)

        if st == "vba":
            wb.remove(ws)
            render_vba_sheet(wb, spec)
            continue

        if st == "kv":
            _title(ws, name)
            r = 3
            for row in (spec.get("rows") or [])[:40]:
                row = list(row)[:3] + [""] * (3 - len(list(row)[:3]))
                ws.cell(row=r, column=1, value=str(row[0])[:60]).font = Font(bold=True, size=10)
                ws.cell(row=r, column=2, value=row[1])
                ws.cell(row=r, column=3, value=str(row[2])[:80]).font = Font(size=9, color=GREY)
                for cc in range(1, 4):
                    ws.cell(row=r, column=cc).border = BORDER
                r += 1
            ws.column_dimensions["A"].width = 34
            ws.column_dimensions["B"].width = 18
            ws.column_dimensions["C"].width = 50

        elif st == "table":
            cols = (spec.get("columns") or [])[:30]
            n = min(int(spec.get("row_count", 20) or 20), 1000)
            hr = 1
            colmap = {}
            for i, c in enumerate(cols, 1):
                cell = ws.cell(row=hr, column=i, value=str(c.get("h", f"Col{i}"))[:40])
                _hdr_cell(cell)
                colmap[c.get("h", f"Col{i}")] = i
                ws.column_dimensions[get_column_letter(i)].width = max(12, min(24, len(str(c.get("h", ""))) + 4))
            ws.row_dimensions[hr].height = 30
            for r in range(hr + 1, hr + 1 + n):
                ctx = {}
                for i, c in enumerate(cols, 1):
                    h = c.get("h", f"Col{i}")
                    if c.get("formula"):
                        v = _translate_formula(c["formula"], colmap, r)
                    else:
                        v = gen.value(c.get("gen") or {"kind": "number"}, r - hr - 1, ctx, salt=i)
                    ctx[h] = v
                    cell = ws.cell(row=r, column=i, value=v)
                    fmt = c.get("format")
                    if fmt:
                        cell.number_format = _fmt_code(fmt, currency_symbol)
                    cell.border = BORDER
                    if (r - hr) % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor=LIGHT)
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{hr + n}"
            # conditional formatting on the first percent column
            for i, c in enumerate(cols, 1):
                if c.get("format") == "percent":
                    cl = get_column_letter(i)
                    ws.conditional_formatting.add(
                        f"{cl}2:{cl}{hr + n}",
                        ColorScaleRule(start_type="min", start_color=RED,
                                       mid_type="percentile", mid_value=50, mid_color="FDE68A",
                                       end_type="max", end_color=GREEN))
                    break
            sheet_meta[name] = {"colmap": colmap, "n_rows": n, "hr": hr,
                                "cols": cols}

        elif st == "summary":
            src = spec.get("source", "")
            meta = sheet_meta.get(src)
            _title(ws, name, f"Aggregated from '{src}' with native SUMIF/COUNTIF/AVERAGEIF")
            r0 = 4
            if not meta:
                ws.cell(row=r0, column=1, value=f"Source sheet '{src}' not found")
                continue
            gb = spec.get("group_by", "")
            gci = meta["colmap"].get(gb)
            # find group values from the source column's gen rule
            gvals = []
            for c in meta["cols"]:
                if c.get("h") == gb and (c.get("gen") or {}).get("values"):
                    gvals = c["gen"]["values"]
            if not gvals:
                gvals = ["Group 1", "Group 2", "Group 3"]
            aggs = (spec.get("aggregates") or [])[:10]
            hdrs = [gb] + [a.get("h", "Agg") for a in aggs]
            for i, h in enumerate(hdrs, 1):
                _hdr_cell(ws.cell(row=r0, column=i, value=h))
                ws.column_dimensions[get_column_letter(i)].width = 20
            gcl = get_column_letter(gci) if gci else "A"
            lastrow = meta["hr"] + meta["n_rows"]
            for j, gv in enumerate(gvals, start=r0 + 1):
                ws.cell(row=j, column=1, value=gv).font = Font(bold=True, size=10)
                ws.cell(row=j, column=1).border = BORDER
                for i, a in enumerate(aggs, 2):
                    kind = a.get("kind", "sum")
                    acol = meta["colmap"].get(a.get("col", ""), gci or 1)
                    acl = get_column_letter(acol)
                    rng = f"'{src}'!{acl}2:{acl}{lastrow}"
                    crit = f"'{src}'!{gcl}2:{gcl}{lastrow},A{j}"
                    if kind == "count":
                        f = f"=COUNTIF('{src}'!{gcl}2:{gcl}{lastrow},A{j})"
                    elif kind == "avg":
                        f = f"=IFERROR(AVERAGEIF({crit},{rng}),0)"
                    else:
                        f = f"=SUMIF({crit},{rng})"
                    cell = ws.cell(row=j, column=i, value=f)
                    if a.get("format"):
                        cell.number_format = _fmt_code(a["format"], currency_symbol)
                    cell.border = BORDER
            tr = r0 + 1 + len(gvals)
            ws.cell(row=tr, column=1, value="TOTAL").font = Font(bold=True, color=NAVY)
            for i, a in enumerate(aggs, 2):
                cl = get_column_letter(i)
                fn = "AVERAGE" if a.get("kind") == "avg" else "SUM"
                cell = ws.cell(row=tr, column=i, value=f"={fn}({cl}{r0+1}:{cl}{tr-1})")
                if a.get("format"):
                    cell.number_format = _fmt_code(a["format"], currency_symbol)
                cell.font = Font(bold=True)
            sheet_meta[name] = {"summary": True, "r0": r0, "n": len(gvals),
                                "hdrs": hdrs}

        elif st == "dashboard":
            ws.sheet_view.showGridLines = False
            _title(ws, bp.get("title", name), "Executive dashboard — all figures computed live via formulas")
            kpis = (spec.get("kpis") or [])[:12]
            r0 = 4
            for i, k in enumerate(kpis):
                col = 1 + (i % 4) * 2
                row = r0 + (i // 4) * 4
                ws.cell(row=row, column=col, value=str(k.get("label", "KPI"))[:30]).font = \
                    Font(bold=True, size=9, color=GREY)
                ref = k.get("ref") or {}
                srcm = sheet_meta.get(ref.get("sheet", ""), {})
                if srcm and not srcm.get("summary"):
                    ci = srcm["colmap"].get(ref.get("col", ""), 1)
                    cl = get_column_letter(ci)
                    lr = srcm["hr"] + srcm["n_rows"]
                    agg = ref.get("agg", "sum").upper()
                    fn = {"COUNT": "COUNTA", "AVG": "AVERAGE", "SUM": "SUM",
                          "MAX": "MAX", "MIN": "MIN"}.get(agg, "SUM")
                    val = f"={fn}('{ref.get('sheet')}'!{cl}2:{cl}{lr})"
                elif k.get("formula"):
                    val = "=" + str(k["formula"]).lstrip("=")
                else:
                    val = k.get("value", "")
                c2 = ws.cell(row=row + 1, column=col, value=val)
                c2.font = Font(bold=True, size=14, color=NAVY)
                if k.get("format"):
                    c2.number_format = _fmt_code(k["format"], currency_symbol)
                for rr in range(row, row + 3):
                    ws.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=LIGHT)
                    ws.cell(row=rr, column=col + 1).fill = PatternFill("solid", fgColor=LIGHT)
            for i in range(1, 9):
                ws.column_dimensions[get_column_letter(i)].width = 16
            # charts from summary sheets
            crow = r0 + ((len(kpis) + 3) // 4) * 4 + 2
            for ch_spec in (spec.get("charts") or [])[:4]:
                srcname = ch_spec.get("source", "")
                sm = sheet_meta.get(srcname, {})
                if not sm.get("summary"):
                    continue
                src_ws = wb[srcname]
                hdrs = sm["hdrs"]
                try:
                    vi = hdrs.index(ch_spec.get("val_col", hdrs[1] if len(hdrs) > 1 else hdrs[0])) + 1
                except ValueError:
                    vi = 2
                ctype = ch_spec.get("type", "bar")
                ch = LineChart() if ctype == "line" else (PieChart() if ctype == "pie" else BarChart())
                if isinstance(ch, BarChart):
                    ch.type = "col"
                ch.title = str(ch_spec.get("title", "Chart"))[:60]
                ch.height, ch.width = 8, 15
                data = Reference(src_ws, min_col=vi, max_col=vi,
                                 min_row=sm["r0"], max_row=sm["r0"] + sm["n"])
                cats = Reference(src_ws, min_col=1, max_col=1,
                                 min_row=sm["r0"] + 1, max_row=sm["r0"] + sm["n"])
                ch.add_data(data, titles_from_data=True)
                ch.set_categories(cats)
                ws.add_chart(ch, f"A{crow}")
                crow += 17

    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def validate_blueprint(bp) -> bool:
    if not isinstance(bp, dict):
        return False
    sheets = bp.get("sheets")
    if not isinstance(sheets, list) or not sheets:
        return False
    has_table = any(s.get("type") == "table" and s.get("columns") for s in sheets
                    if isinstance(s, dict))
    return has_table


# ── v4.2: VBA/Automation sheet type ──────────────────────────────
# openpyxl cannot embed a compiled vbaProject.bin into a new workbook, so
# automation requests get an "Automation (VBA)" sheet carrying ready-to-paste
# VBA modules + install instructions. Honest, works everywhere, no corruption.
def render_vba_sheet(wb, spec):
    from openpyxl.styles import Font, PatternFill, Border
    ws = wb.create_sheet(str(spec.get("name", "Automation (VBA)"))[:28])
    ws["A1"] = "Automation — VBA modules (copy into Alt+F11 → Insert → Module)"
    ws["A1"].font = Font(bold=True, size=13, color="1E3A5F")
    ws["A2"] = "Then save the file as .xlsm (macro-enabled). Each module below is complete and ready to run."
    ws["A2"].font = Font(size=9, color="64748B", italic=True)
    r = 4
    for m in (spec.get("modules") or [])[:6]:
        ws.cell(row=r, column=1, value="Module: " + str(m.get("name", "Macro"))[:50]).font = \
            Font(bold=True, size=11, color="0F6E56")
        r += 1
        for line in str(m.get("code", ""))[:6000].split("\n")[:120]:
            c = ws.cell(row=r, column=1, value=line)
            c.font = Font(name="Consolas", size=9)
            r += 1
        r += 2
    ws.column_dimensions["A"].width = 110
    return ws
