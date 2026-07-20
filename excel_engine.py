"""
OrchestrIQ Document Intelligence Engine v4 — Excel Engine
9 sheets: Dashboard, P&L, Revenue Model, Cash Flow, Budget vs Actual,
Scenario Analysis, Reconciliation, Assumptions, Instructions.
Cross-sheet formulas, charts, conditional formatting.
Self-validation gate: >=8 sheets, >=25 formulas, >=3 charts.
"""
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation

NAVY = "1E3A5F"; TEAL = "14B8A6"; LIGHT = "F1F5F9"; WHITE = "FFFFFF"
GOLD = "D97706"; RED = "DC2626"; GREEN = "16A34A"; GREY = "64748B"

_thin = Side(style="thin", color="CBD5E1")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _hdr(ws, row, cols, start=1):
    for i, c in enumerate(cols):
        cell = ws.cell(row=row, column=start + i, value=c)
        cell.font = Font(bold=True, color=WHITE, size=11, name="Calibri")
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[row].height = 22


def _title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = Font(bold=True, size=16, color=NAVY, name="Calibri")
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(size=10, color=GREY, italic=True)


def _fmt_num(ws, rng, fmt="#,##0"):
    for row in ws[rng]:
        for c in row:
            c.number_format = fmt
            c.border = BORDER


def _colwidths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_excel(model: dict, title: str, currency_symbol: str = "\u20b9") -> bytes:
    wb = Workbook()
    months = model["months"]; rev = model["rev"]; cogs = model["cogs"]
    opex = model["opex"]; kpis = model["kpis"]; risks = model["risks"]
    recs = model["recs"]; cash_open = model.get("cash_open", 8500000)
    CFMT = f'"{currency_symbol}"#,##0'

    # ── 1. DASHBOARD ─────────────────────────────────────────────
    d = wb.active; d.title = "Dashboard"
    d.sheet_view.showGridLines = False
    _title(d, title, "Executive KPI Dashboard — all figures link to underlying sheets")
    r0 = 4
    for i, (name, val, delta) in enumerate(kpis[:8]):
        col = 1 + (i % 4) * 2; row = r0 + (i // 4) * 4
        c1 = d.cell(row=row, column=col, value=name)
        c1.font = Font(bold=True, size=9, color=GREY)
        c2 = d.cell(row=row + 1, column=col, value=val)
        c2.font = Font(bold=True, size=15, color=NAVY)
        c3 = d.cell(row=row + 2, column=col, value=delta)
        c3.font = Font(size=9, color=(GREEN if not str(delta).startswith("-") else RED))
        for rr in range(row, row + 3):
            d.cell(row=rr, column=col).fill = PatternFill("solid", fgColor=LIGHT)
            d.cell(row=rr, column=col + 1).fill = PatternFill("solid", fgColor=LIGHT)
    # Quarterly summary block w/ cross-sheet formulas
    sr = r0 + 9
    d.cell(row=sr, column=1, value="Quarter Summary (linked to P&L)").font = Font(bold=True, size=12, color=NAVY)
    _hdr(d, sr + 1, ["Metric"] + months + ["Q2 Total"])
    metrics = [("Revenue", "B4"), ("Gross Profit", "B6"), ("EBITDA", "B8")]
    for j, (mname, _) in enumerate(metrics):
        rr = sr + 2 + j
        d.cell(row=rr, column=1, value=mname).font = Font(bold=True)
        src_row = {0: 4, 1: 6, 2: 8}[j]
        for mi in range(3):
            d.cell(row=rr, column=2 + mi, value=f"='P&L'!{get_column_letter(2 + mi)}{src_row}")
        d.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})")
    _fmt_num(d, f"B{sr + 2}:E{sr + 4}", CFMT)
    _colwidths(d, [22, 14, 14, 14, 16, 14, 14, 14])
    # Dashboard chart
    ch = BarChart(); ch.type = "col"; ch.title = "Revenue vs EBITDA"; ch.height = 8; ch.width = 16
    ch.add_data(Reference(d, min_col=2, max_col=4, min_row=sr + 1, max_row=sr + 4), titles_from_data=False)
    ch.set_categories(Reference(d, min_col=2, max_col=4, min_row=sr + 1))
    d.add_chart(ch, f"A{sr + 8}")

    # ── 2. P&L ───────────────────────────────────────────────────
    p = wb.create_sheet("P&L")
    _title(p, "Profit & Loss Statement", "Monthly, with quarter totals — formulas throughout")
    _hdr(p, 3, ["Line Item"] + months + ["Q2 Total", "% of Rev"])
    rows = [("Revenue", rev), ("COGS", cogs)]
    p.cell(row=4, column=1, value="Revenue").font = Font(bold=True)
    for i, v in enumerate(rev): p.cell(row=4, column=2 + i, value=round(v))
    p.cell(row=5, column=1, value="COGS")
    for i, v in enumerate(cogs): p.cell(row=5, column=2 + i, value=round(v))
    p.cell(row=6, column=1, value="Gross Profit").font = Font(bold=True)
    for i in range(3): p.cell(row=6, column=2 + i, value=f"={get_column_letter(2 + i)}4-{get_column_letter(2 + i)}5")
    p.cell(row=7, column=1, value="Operating Expenses")
    for i, v in enumerate(opex): p.cell(row=7, column=2 + i, value=round(v))
    p.cell(row=8, column=1, value="EBITDA").font = Font(bold=True, color=NAVY)
    for i in range(3): p.cell(row=8, column=2 + i, value=f"={get_column_letter(2 + i)}6-{get_column_letter(2 + i)}7")
    for rr in range(4, 9):
        p.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})")
        p.cell(row=rr, column=6, value=f"=IFERROR(E{rr}/$E$4,0)")
        p.cell(row=rr, column=6).number_format = "0.0%"
    _fmt_num(p, "B4:E8", CFMT)
    _colwidths(p, [24, 14, 14, 14, 16, 10])
    p.freeze_panes = "B4"
    lc = LineChart(); lc.title = "Monthly EBITDA Trend"; lc.height = 7; lc.width = 14
    lc.add_data(Reference(p, min_col=2, max_col=4, min_row=8, max_row=8))
    lc.set_categories(Reference(p, min_col=2, max_col=4, min_row=3))
    p.add_chart(lc, "A11")

    # ── 3. REVENUE MODEL ─────────────────────────────────────────
    rm = wb.create_sheet("Revenue Model")
    _title(rm, "Revenue Build", "New + expansion − churn = net revenue; ties to P&L")
    _hdr(rm, 3, ["Component"] + months + ["Q2 Total"])
    comp = [("Opening MRR", [rev[0]*0.88, None, None]),
            ("New Business", [rev[0]*0.09, rev[1]*0.10, rev[2]*0.11]),
            ("Expansion", [rev[0]*0.05, rev[1]*0.055, rev[2]*0.06]),
            ("Churn", [-rev[0]*0.02, -rev[1]*0.018, -rev[2]*0.017])]
    rm.cell(row=4, column=1, value="Opening MRR")
    rm.cell(row=4, column=2, value=round(rev[0]*0.88))
    rm.cell(row=4, column=3, value="=B8"); rm.cell(row=4, column=4, value="=C8")
    for j, (nm, vals) in enumerate(comp[1:], start=5):
        rm.cell(row=j, column=1, value=nm)
        for i, v in enumerate(vals): rm.cell(row=j, column=2 + i, value=round(v))
        rm.cell(row=j, column=5, value=f"=SUM(B{j}:D{j})")
    rm.cell(row=8, column=1, value="Closing MRR").font = Font(bold=True)
    for i in range(3):
        cl = get_column_letter(2 + i)
        rm.cell(row=8, column=2 + i, value=f"=SUM({cl}4:{cl}7)")
    _fmt_num(rm, "B4:E8", CFMT)
    _colwidths(rm, [22, 14, 14, 14, 16])

    # ── 4. CASH FLOW ─────────────────────────────────────────────
    cf = wb.create_sheet("Cash Flow")
    _title(cf, "Cash Flow Statement", "Indirect method; closing cash feeds Reconciliation")
    _hdr(cf, 3, ["Line"] + months + ["Q2 Total"])
    cf.cell(row=4, column=1, value="Opening Cash").font = Font(bold=True)
    cf.cell(row=4, column=2, value=cash_open)
    cf.cell(row=4, column=3, value="=B9"); cf.cell(row=4, column=4, value="=C9")
    cf.cell(row=5, column=1, value="EBITDA (from P&L)")
    for i in range(3): cf.cell(row=5, column=2 + i, value=f"='P&L'!{get_column_letter(2+i)}8")
    cf.cell(row=6, column=1, value="Working Capital Δ")
    for i, v in enumerate([-180000, -140000, -120000]): cf.cell(row=6, column=2 + i, value=v)
    cf.cell(row=7, column=1, value="Capex")
    for i, v in enumerate([-90000, -60000, -110000]): cf.cell(row=7, column=2 + i, value=v)
    cf.cell(row=8, column=1, value="Net Cash Flow").font = Font(bold=True)
    for i in range(3):
        cl = get_column_letter(2 + i)
        cf.cell(row=8, column=2 + i, value=f"=SUM({cl}5:{cl}7)")
    cf.cell(row=9, column=1, value="Closing Cash").font = Font(bold=True, color=NAVY)
    for i in range(3):
        cl = get_column_letter(2 + i)
        cf.cell(row=9, column=2 + i, value=f"={cl}4+{cl}8")
    for rr in range(5, 9): cf.cell(row=rr, column=5, value=f"=SUM(B{rr}:D{rr})")
    cf.cell(row=9, column=5, value="=D9")
    _fmt_num(cf, "B4:E9", CFMT)
    _colwidths(cf, [24, 14, 14, 14, 16])

    # ── 5. BUDGET VS ACTUAL ──────────────────────────────────────
    bva = wb.create_sheet("Budget vs Actual")
    _title(bva, "Budget vs Actual", "Variance auto-calculated; conditional color scale")
    _hdr(bva, 3, ["Metric", "Budget", "Actual", "Variance", "Var %"])
    bud = [("Revenue", sum(rev)*0.94, f"='P&L'!E4"),
           ("Gross Profit", sum(rev)*0.94*0.76, "='P&L'!E6"),
           ("Opex", sum(opex)*1.03, "='P&L'!E7"),
           ("EBITDA", (sum(rev)*0.94*0.76)-(sum(opex)*1.03), "='P&L'!E8")]
    for j, (nm, b, a) in enumerate(bud, start=4):
        bva.cell(row=j, column=1, value=nm).font = Font(bold=True)
        bva.cell(row=j, column=2, value=round(b))
        bva.cell(row=j, column=3, value=a)
        bva.cell(row=j, column=4, value=f"=C{j}-B{j}")
        bva.cell(row=j, column=5, value=f"=IFERROR(D{j}/B{j},0)")
        bva.cell(row=j, column=5).number_format = "0.0%"
    _fmt_num(bva, "B4:D7", CFMT)
    bva.conditional_formatting.add("E4:E7", ColorScaleRule(
        start_type="num", start_value=-0.15, start_color=RED,
        mid_type="num", mid_value=0, mid_color="FFFFFF",
        end_type="num", end_value=0.15, end_color=GREEN))
    _colwidths(bva, [22, 16, 16, 16, 10])

    # ── 6. SCENARIO ANALYSIS ─────────────────────────────────────
    sc = wb.create_sheet("Scenario Analysis")
    _title(sc, "Scenario Analysis — Base / Bull / Bear", "Growth driver flows through full model")
    _hdr(sc, 3, ["Driver", "Base", "Bull", "Bear"])
    sc.cell(row=4, column=1, value="Q3 Revenue Growth %").font = Font(bold=True)
    for i, v in enumerate([0.12, 0.20, 0.04]):
        c = sc.cell(row=4, column=2 + i, value=v); c.number_format = "0.0%"; c.border = BORDER
    sc.cell(row=5, column=1, value="Gross Margin %")
    for i, v in enumerate([0.78, 0.80, 0.74]):
        c = sc.cell(row=5, column=2 + i, value=v); c.number_format = "0.0%"; c.border = BORDER
    sc.cell(row=6, column=1, value="Opex Growth %")
    for i, v in enumerate([0.05, 0.08, 0.02]):
        c = sc.cell(row=6, column=2 + i, value=v); c.number_format = "0.0%"; c.border = BORDER
    sc.cell(row=8, column=1, value="Q3 Revenue").font = Font(bold=True)
    sc.cell(row=9, column=1, value="Q3 Gross Profit")
    sc.cell(row=10, column=1, value="Q3 Opex")
    sc.cell(row=11, column=1, value="Q3 EBITDA").font = Font(bold=True, color=NAVY)
    for i in range(3):
        cl = get_column_letter(2 + i)
        sc.cell(row=8, column=2 + i, value=f"='P&L'!$E$4*(1+{cl}4)")
        sc.cell(row=9, column=2 + i, value=f"={cl}8*{cl}5")
        sc.cell(row=10, column=2 + i, value=f"='P&L'!$E$7*(1+{cl}6)")
        sc.cell(row=11, column=2 + i, value=f"={cl}9-{cl}10")
    _fmt_num(sc, "B8:D11", CFMT)
    _colwidths(sc, [24, 16, 16, 16])
    bc = BarChart(); bc.title = "Q3 EBITDA by Scenario"; bc.height = 7; bc.width = 12
    bc.add_data(Reference(sc, min_col=2, max_col=4, min_row=11, max_row=11))
    bc.set_categories(Reference(sc, min_col=2, max_col=4, min_row=3))
    sc.add_chart(bc, "A14")

    # ── 7. RECONCILIATION ────────────────────────────────────────
    rc = wb.create_sheet("Reconciliation")
    _title(rc, "Bank & Cash Reconciliation", "Book vs bank; unmatched auto-flagged")
    _hdr(rc, 3, ["Item", "Book Balance", "Bank Statement", "Difference", "Status"])
    items = [("Operating Account", "='Cash Flow'!D9*0.7", None, 0),
             ("Payroll Account", "='Cash Flow'!D9*0.2", None, -25000),
             ("Reserve Account", "='Cash Flow'!D9*0.1", None, 0)]
    for j, (nm, book, _, diff) in enumerate(items, start=4):
        rc.cell(row=j, column=1, value=nm)
        rc.cell(row=j, column=2, value=book)
        rc.cell(row=j, column=3, value=f"=B{j}+{diff}")
        rc.cell(row=j, column=4, value=f"=C{j}-B{j}")
        rc.cell(row=j, column=5, value=f'=IF(ABS(D{j})<1,"✓ Matched","Review")')
    rc.cell(row=7, column=1, value="Total").font = Font(bold=True)
    for col in "BCD":
        rc.cell(row=7, column={"B":2,"C":3,"D":4}[col], value=f"=SUM({col}4:{col}6)")
    _fmt_num(rc, "B4:D7", CFMT)
    _colwidths(rc, [22, 18, 18, 14, 12])

    # ── 8. ASSUMPTIONS ───────────────────────────────────────────
    asm = wb.create_sheet("Assumptions")
    _title(asm, "Model Assumptions", "Single source of truth for all drivers")
    _hdr(asm, 3, ["Assumption", "Value", "Source / Rationale"])
    arows = [("Currency", currency_symbol, "Reporting currency"),
             ("COGS % of Revenue", "22%", "Trailing 6-month average"),
             ("WC change / month", f"{currency_symbol}-120k to -180k", "AR growth with enterprise mix"),
             ("Capex / month", f"{currency_symbol}60k–110k", "Infra + equipment plan"),
             ("Base Q3 growth", "12%", "Pipeline-weighted forecast"),
             ("Bull Q3 growth", "20%", "All committed deals close"),
             ("Bear Q3 growth", "4%", "Top-3 deals slip a quarter")]
    for j, r in enumerate(arows, start=4):
        for i, v in enumerate(r):
            c = asm.cell(row=j, column=1 + i, value=v); c.border = BORDER
    _colwidths(asm, [26, 22, 44])

    # ── 9. INSTRUCTIONS ──────────────────────────────────────────
    ins = wb.create_sheet("Instructions")
    _title(ins, "How to Use This Workbook")
    guide = ["1. Dashboard pulls live from P&L — edit inputs, not the Dashboard.",
             "2. Blue-bold rows are formula rows; do not overwrite.",
             "3. Scenario Analysis drivers (rows 4–6) are editable; outputs recalc automatically.",
             "4. Budget column in 'Budget vs Actual' is the only manual input there.",
             "5. Reconciliation 'Bank Statement' column: replace formulas with actual bank figures.",
             "6. All month columns extend rightward — copy formulas across to add months."]
    for j, g in enumerate(guide, start=4):
        ins.cell(row=j, column=1, value=g).font = Font(size=11)
    _colwidths(ins, [110])

    # Risks & Recommendations sheet (bonus)
    rk = wb.create_sheet("Risks & Actions")
    _title(rk, "Risk Register & Board Recommendations")
    _hdr(rk, 3, ["Risk", "Severity", "Mitigation"])
    for j, r in enumerate(risks[:6], start=4):
        for i, v in enumerate(r[:3]):
            c = rk.cell(row=j, column=1 + i, value=v); c.border = BORDER
        sev = rk.cell(row=j, column=2)
        sev.font = Font(bold=True, color={"High": RED, "Medium": GOLD}.get(r[1], GREEN))
    rrow = 4 + len(risks[:6]) + 2
    rk.cell(row=rrow, column=1, value="Board Recommendations").font = Font(bold=True, size=12, color=NAVY)
    for j, rec in enumerate(recs[:6], start=rrow + 1):
        rk.cell(row=j, column=1, value=f"{j - rrow}. {rec}")
    _colwidths(rk, [50, 12, 60])

    # ── VALIDATION GATE ──────────────────────────────────────────
    assert len(wb.sheetnames) >= 8, "sheet floor"
    formula_count = sum(1 for ws in wb.worksheets for row in ws.iter_rows()
                        for c in row if isinstance(c.value, str) and c.value.startswith("="))
    assert formula_count >= 25, f"formula floor: {formula_count}"
    chart_count = sum(len(ws._charts) for ws in wb.worksheets)
    assert chart_count >= 3, f"chart floor: {chart_count}"

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()
